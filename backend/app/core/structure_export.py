"""Экспорт структуры данных в XLSX с цветовой маркировкой типов."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
KIND_LABELS = {
    "numeric": "Число",
    "categorical": "Категория",
    "datetime": "Дата",
    "boolean": "Булев",
    "identifier": "ID",
    "textual": "Текст",
}

KIND_FILLS = {
    "numeric": ("DBEAFE", "1D4ED8"),
    "categorical": ("EDE9FE", "6D28D9"),
    "datetime": ("D1FAE5", "047857"),
    "boolean": ("FEF3C7", "B45309"),
    "identifier": ("F3F4F6", "4B5563"),
    "textual": ("FCE7F3", "BE185D"),
}

HEADER_FILL = PatternFill("solid", fgColor="F3F4F6")
HEADER_FONT = Font(bold=True, color="4B5563", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)


def _format_stats(description: str) -> str:
    text = description or ""
    filled = re.search(r"(\d+)\s+значений?", text)
    unique = re.search(r"(\d+)\s+(уникальных|категорий)", text)
    parts: list[str] = []
    if filled:
        parts.append(f"{filled.group(1)} зап.")
    if unique:
        suffix = "кат." if unique.group(2) == "категорий" else "уник."
        parts.append(f"{unique.group(1)} {suffix}")
    return " · ".join(parts) if parts else "—"


def build_structure_xlsx(data_structure: dict, output_path: Path) -> None:
    columns = data_structure.get("columns", [])
    datetime_candidates = data_structure.get("datetime_candidates", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Структура"

    row = 1
    ws.cell(row, 1, "Структура данных").font = Font(bold=True, size=14, color="111827")
    row += 1
    ws.cell(row, 1, f"Столбцов: {len(columns)}").font = Font(size=11, color="6B7280")
    row += 1
    if datetime_candidates:
        ws.cell(row, 1, f"Столбцы с датой: {', '.join(datetime_candidates)}").font = Font(
            size=10, color="6B7280", italic=True
        )
        row += 1
    row += 1

    headers = ["#", "Столбец", "Тип", "Сводка"]
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row, col_idx, title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left", vertical="center")

    header_row = row
    row += 1

    for i, col in enumerate(columns, start=1):
        kind = col.get("kind") or "textual"
        kind_label = KIND_LABELS.get(kind, col.get("type", kind))
        stats = _format_stats(col.get("description", ""))

        ws.cell(row, 1, i).alignment = Alignment(horizontal="center")
        ws.cell(row, 1).font = Font(color="9CA3AF", size=10)
        ws.cell(row, 2, col.get("name", "")).font = Font(bold=True, color="111827")
        type_cell = ws.cell(row, 3, kind_label)
        bg, fg = KIND_FILLS.get(kind, KIND_FILLS["textual"])
        type_cell.fill = PatternFill("solid", fgColor=bg)
        type_cell.font = Font(bold=True, color=fg, size=10)
        type_cell.alignment = Alignment(horizontal="center")
        ws.cell(row, 4, stats).font = Font(color="6B7280", size=10)

        for col_idx in range(1, 5):
            ws.cell(row, col_idx).border = THIN_BORDER
            if col_idx != 1:
                ws.cell(row, col_idx).alignment = Alignment(vertical="center")

        row += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.freeze_panes = f"A{header_row + 1}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
