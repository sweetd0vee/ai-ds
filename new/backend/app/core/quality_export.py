"""Экспорт качества данных и связей между столбцами (TXT, XLSX)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .data_insights import CORRELATION_HELP
from .structure_export import HEADER_FILL, HEADER_FONT, KIND_LABELS, THIN_BORDER

ISSUE_LABELS = {
    "high_missing": "Много пропусков (>50%)",
    "moderate_missing": "Есть пропуски (10–50%)",
    "constant": "Константа",
    "near_unique": "Почти все значения уникальны",
    "likely_identifier": "Похоже на ID",
}

STRENGTH_FILLS = {
    "сильная": ("DBEAFE", "1D4ED8"),
    "умеренная": ("E0E7FF", "4338CA"),
}


def _issues_text(issues: list[str]) -> str:
    return "; ".join(ISSUE_LABELS.get(i, i) for i in issues) if issues else "—"


def format_insights_report(quality: dict, correlations: dict) -> str:
    """Единый человекочитаемый отчёт для копирования и TXT-скачивания."""
    summary = quality.get("summary", {})
    kinds = summary.get("column_kinds") or {}
    kind_line = ", ".join(
        f"{KIND_LABELS.get(k, k)}: {v}" for k, v in sorted(kinds.items(), key=lambda x: -x[1])
    )

    lines = [
        "КАЧЕСТВО ДАННЫХ И СВЯЗИ МЕЖДУ СТОЛБЦАМИ",
        "=" * 44,
        "",
        "── СВОДКА ──",
        f"Оценка качества: {summary.get('overall_score', '—')}/100 ({summary.get('overall_grade_label', '—')})",
        f"Размер: {summary.get('rows', 0)} строк × {summary.get('columns', 0)} столбцов",
        f"Заполненность ячеек: {summary.get('fill_rate_pct', 0)}%",
        f"Строк без пропусков: {summary.get('complete_rows', 0)} ({summary.get('complete_rows_pct', 0)}%)",
        f"Дубликаты строк: {summary.get('duplicate_rows', 0)} ({summary.get('duplicate_pct', 0)}%)",
        f"Средний % пропусков по столбцам: {summary.get('avg_missing_pct', 0)}%",
        f"Столбцов для анализа (без ID и констант): {summary.get('usable_columns', 0)}",
        f"Состав: {kind_line or '—'}",
        f"Идентификаторов: {summary.get('identifier_columns', 0)}, "
        f"констант: {summary.get('constant_columns', 0)}, "
        f"с >50% пропусков: {summary.get('columns_with_high_missing', 0)}, "
        f"с умеренными пропусками: {summary.get('moderate_missing_columns', 0)}",
        "",
    ]

    top_missing = summary.get("top_missing_columns") or []
    if top_missing:
        lines.extend(["── ТОП ПРОПУСКОВ ──"])
        for col in top_missing:
            lines.append(
                f"  • {col['name']} ({KIND_LABELS.get(col.get('kind'), col.get('kind', ''))}): "
                f"{col['missing_pct']}%"
            )
        lines.append("")

    flagged = [c for c in quality.get("columns", []) if c.get("issues")]
    lines.append("── ЗАМЕЧАНИЯ ПО СТОЛБЦАМ ──")
    if flagged:
        for col in flagged:
            lines.append(
                f"  • {col['name']} | {KIND_LABELS.get(col['kind'], col['kind'])} | "
                f"пропуски {col['missing_pct']}% | уникальных {col['nunique']} | "
                f"{_issues_text(col['issues'])}"
            )
    else:
        lines.append("  Критичных проблем не обнаружено.")
    lines.append("")

    lines.extend([
        "── СПРАВКА ПО КОЭФФИЦИЕНТАМ ──",
        correlations.get("filter_note", "Показаны только сильные и умеренные связи."),
        "",
    ])
    for key, help_item in (correlations.get("help") or CORRELATION_HELP).items():
        lines.extend([
            f"{help_item['name']} ({help_item['range']})",
            f"  {help_item['meaning']}",
            f"  Пороги: {help_item['thresholds']}",
            "",
        ])

    lines.append("── СВЯЗИ МЕЖДУ СТОЛБЦАМИ ──")
    num_pairs = correlations.get("numeric_pairs") or []
    if num_pairs:
        lines.append("")
        lines.append("Числовые (Pearson r):")
        for p in num_pairs:
            lines.append(
                f"  • {p['col_a']} ↔ {p['col_b']}: r={p['pearson']} | "
                f"{p.get('direction', '')} | {p['strength']}"
            )
    else:
        lines.append("  Числовые: сильных/умеренных связей не найдено.")

    cat_pairs = correlations.get("categorical_pairs") or []
    if cat_pairs:
        lines.append("")
        lines.append("Категориальные (Cramér's V):")
        for p in cat_pairs:
            lines.append(f"  • {p['col_a']} ↔ {p['col_b']}: V={p['cramers_v']} | {p['strength']}")

    cat_num = correlations.get("categorical_numeric") or []
    if cat_num:
        lines.append("")
        lines.append("Категория → число (η):")
        for p in cat_num:
            lines.append(
                f"  • {p['categorical']} → {p['numeric']}: η={p['eta']} | {p['strength']}"
            )

    if not num_pairs and not cat_pairs and not cat_num:
        lines.append("  Значимых связей не обнаружено.")

    return "\n".join(lines)


def _write_header_row(ws, row: int, headers: list[str]) -> int:
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row, col_idx, title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left", vertical="center")
    return row + 1


def _style_strength_cell(cell, strength: str) -> None:
    bg, fg = STRENGTH_FILLS.get(strength, ("F3F4F6", "4B5563"))
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=fg, size=10)
    cell.alignment = Alignment(horizontal="center")


def _autosize_columns(ws, widths: dict[str, float]) -> None:
    from openpyxl.utils import get_column_letter

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def build_quality_xlsx(
    quality: dict,
    correlations: dict,
    output_path: Path,
    *,
    source_file: str = "",
) -> None:
    """XLSX: сводка, столбцы, пропуски, корреляции, справка."""
    summary = quality.get("summary", {})
    wb = Workbook()

    # ── Лист: Сводка ──
    ws_sum = wb.active
    ws_sum.title = "Сводка"
    r = 1
    ws_sum.cell(r, 1, "Качество данных и связи").font = Font(bold=True, size=14)
    r += 1
    if source_file:
        ws_sum.cell(r, 1, f"Файл: {source_file}").font = Font(size=10, color="6B7280")
        r += 1
    r += 1

    metrics = [
        ("Оценка качества", f"{summary.get('overall_score', '—')}/100"),
        ("Уровень", summary.get("overall_grade_label", "—")),
        ("Строк", summary.get("rows", 0)),
        ("Столбцов", summary.get("columns", 0)),
        ("Заполненность ячеек", f"{summary.get('fill_rate_pct', 0)}%"),
        ("Строк без пропусков", f"{summary.get('complete_rows', 0)} ({summary.get('complete_rows_pct', 0)}%)"),
        ("Дубликаты строк", f"{summary.get('duplicate_rows', 0)} ({summary.get('duplicate_pct', 0)}%)"),
        ("Средний % пропусков", f"{summary.get('avg_missing_pct', 0)}%"),
        ("Столбцов для анализа", summary.get("usable_columns", 0)),
        ("Идентификаторов", summary.get("identifier_columns", 0)),
        ("Констант", summary.get("constant_columns", 0)),
        (">50% пропусков", summary.get("columns_with_high_missing", 0)),
        ("Умеренные пропуски", summary.get("moderate_missing_columns", 0)),
    ]
    for label, value in metrics:
        ws_sum.cell(r, 1, label).font = Font(bold=True, color="374151")
        ws_sum.cell(r, 2, value)
        r += 1

    kinds = summary.get("column_kinds") or {}
    if kinds:
        r += 1
        ws_sum.cell(r, 1, "Состав по типам").font = Font(bold=True, size=11)
        r += 1
        for kind, count in sorted(kinds.items(), key=lambda x: -x[1]):
            ws_sum.cell(r, 1, KIND_LABELS.get(kind, kind))
            ws_sum.cell(r, 2, count)
            r += 1

    _autosize_columns(ws_sum, {1: 28, 2: 22})

    # ── Лист: Столбцы ──
    ws_cols = wb.create_sheet("Столбцы")
    r = _write_header_row(
        ws_cols, 1,
        ["Столбец", "Тип", "Заполнено", "Пропуски", "Пропуски %", "Уникальных", "Уникальных %", "Замечания"],
    )
    for col in quality.get("columns", []):
        ws_cols.cell(r, 1, col.get("name", ""))
        ws_cols.cell(r, 2, KIND_LABELS.get(col.get("kind"), col.get("kind", "")))
        ws_cols.cell(r, 3, col.get("non_null", 0))
        ws_cols.cell(r, 4, col.get("missing", 0))
        ws_cols.cell(r, 5, col.get("missing_pct", 0))
        ws_cols.cell(r, 6, col.get("nunique", 0))
        ws_cols.cell(r, 7, col.get("unique_pct", 0))
        ws_cols.cell(r, 8, _issues_text(col.get("issues", [])))
        for c in range(1, 9):
            ws_cols.cell(r, c).border = THIN_BORDER
        r += 1
    ws_cols.freeze_panes = "A2"
    _autosize_columns(ws_cols, {1: 24, 2: 14, 3: 12, 4: 10, 5: 12, 6: 12, 7: 14, 8: 36})

    # ── Лист: Пропуски ──
    ws_miss = wb.create_sheet("Пропуски")
    r = _write_header_row(ws_miss, 1, ["Столбец", "Тип", "Пропуски %", "Пропущено", "Заполнено"])
    missing_cols = sorted(
        quality.get("columns", []),
        key=lambda c: c.get("missing_pct", 0),
        reverse=True,
    )
    missing_cols = [c for c in missing_cols if c.get("missing_pct", 0) > 0]
    if not missing_cols:
        ws_miss.cell(2, 1, "Пропусков нет")
    else:
        r = 2
        for col in missing_cols:
            ws_miss.cell(r, 1, col.get("name", ""))
            ws_miss.cell(r, 2, KIND_LABELS.get(col.get("kind"), col.get("kind", "")))
            ws_miss.cell(r, 3, col.get("missing_pct", 0))
            ws_miss.cell(r, 4, col.get("missing", 0))
            ws_miss.cell(r, 5, col.get("non_null", 0))
            for c in range(1, 6):
                ws_miss.cell(r, c).border = THIN_BORDER
            r += 1
    ws_miss.freeze_panes = "A2"
    _autosize_columns(ws_miss, {1: 24, 2: 14, 3: 12, 4: 12, 5: 12})

    # ── Листы корреляций ──
    def _corr_sheet(title: str, headers: list[str], rows: list[list], strength_col: int | None) -> None:
        ws = wb.create_sheet(title)
        r = _write_header_row(ws, 1, headers)
        if not rows:
            ws.cell(2, 1, "Сильных и умеренных связей не найдено")
            return
        for row_data in rows:
            for c, val in enumerate(row_data, start=1):
                cell = ws.cell(r, c, val)
                cell.border = THIN_BORDER
                if strength_col and c == strength_col and isinstance(val, str):
                    _style_strength_cell(cell, val)
            r += 1
        ws.freeze_panes = "A2"

    num_rows = [
        [p["col_a"], p["col_b"], p["pearson"], p.get("direction", ""), p["strength"]]
        for p in correlations.get("numeric_pairs") or []
    ]
    _corr_sheet("Корр_Pearson", ["Столбец A", "Столбец B", "r", "Направление", "Сила"], num_rows, 5)

    cat_rows = [
        [p["col_a"], p["col_b"], p["cramers_v"], p["strength"]]
        for p in correlations.get("categorical_pairs") or []
    ]
    _corr_sheet("Корр_Cramers", ["Столбец A", "Столбец B", "V", "Сила"], cat_rows, 4)

    eta_rows = [
        [p["categorical"], p["numeric"], p["eta"], p["strength"]]
        for p in correlations.get("categorical_numeric") or []
    ]
    _corr_sheet("Корр_Eta", ["Категория", "Число", "η", "Сила"], eta_rows, 4)

    # ── Лист: Справка ──
    ws_help = wb.create_sheet("Справка")
    ws_help.cell(1, 1, "Коэффициенты связи").font = Font(bold=True, size=12)
    ws_help.cell(2, 1, correlations.get("filter_note", "")).font = Font(italic=True, color="6B7280")
    r = 4
    for help_item in (correlations.get("help") or CORRELATION_HELP).values():
        ws_help.cell(r, 1, help_item["name"]).font = Font(bold=True)
        ws_help.cell(r, 2, help_item["range"])
        r += 1
        ws_help.cell(r, 1, help_item["meaning"])
        ws_help.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
        ws_help.cell(r, 1, help_item["thresholds"]).font = Font(color="4338CA")
        ws_help.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 2
    _autosize_columns(ws_help, {1: 22, 2: 14, 3: 50})

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
